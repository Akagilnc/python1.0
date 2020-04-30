from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl import load_workbook
from openpyxl.chart.label import DataLabelList
import pandas as pd
from matplotlib import pyplot as plt


# 创建chart
def draw_chart(input_wb, max_row):
    # 拿到我们需要操作的sheet top10
    top_10_sheet = input_wb.worksheets[-1]
    top_10_sheet.sheet_view.zoomScale = 200
    # 初始化chart
    bar_chart = BarChart()
    line_chart = LineChart()
    pie_chart = PieChart()
    # 生成数据
    bar_chart_data = Reference(top_10_sheet, min_col=4, max_col=7, min_row=1, max_row=max_row)
    line_chart_data = Reference(top_10_sheet, min_col=4, max_col=4, min_row=1, max_row=max_row)
    pie_chart_data = Reference(top_10_sheet, min_col=4, max_col=4, min_row=1, max_row=max_row)
    # 指定chart的x_axis
    x_data = Reference(top_10_sheet, min_col=2, max_col=2, min_row=2, max_row=max_row)
    # 设置chart样式
    bar_chart.height, bar_chart.width = 7, 15
    line_chart.height, line_chart.width = 7, 15
    pie_chart.height, pie_chart.width = 7, 15
    bar_chart.title, bar_chart.y_axis.title, bar_chart.x_axis.title = 'top10', '人数', '国家'
    # bar_chart.y_axis.scaling.max = 5000000
    # 把数据添加进chart
    bar_chart.add_data(bar_chart_data, titles_from_data=True)
    line_chart.add_data(line_chart_data, titles_from_data=True)
    bar_chart.set_categories(x_data)
    line_chart.set_categories(x_data)
    pie_chart.add_data(pie_chart_data, titles_from_data=True)
    pie_chart.set_categories(x_data)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showLegendKey = True
    # 把chart添加到sheet
    # top_10_sheet.add_chart(bar_chart, 'I1')
    # top_10_sheet.add_chart(pie_chart, 'I11')
    bar_chart += line_chart
    top_10_sheet.add_chart(bar_chart, 'I1')
    # 保存我们的workbook
    input_wb.save('./excel_files/report_chart.xlsx')


# wb = load_workbook('./excel_files/report_formatted.xlsx')
# draw_chart(wb, wb.worksheets[-1].max_row)
plt.rcParams['font.sans-serif'] = ['Simhei']

df = pd.read_excel('./excel_files/report_formatted.xlsx', sheet_name='top10')
df.plot(kind='bar', x='countryName', y=['当前确诊', '累计确诊', '治愈', '死亡'])
# df.set_index(df['countryName']).plot.pie(y='当前确诊')
df.plot(kind='line', x='countryName', y=['当前确诊', '累计确诊', '治愈', '死亡'], secondary_y=True)
ax = plt.gca()
plt.show()
