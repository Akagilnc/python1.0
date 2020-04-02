from openpyxl import Workbook
import requests
import json
import pandas as pd


# 获取最新数据 
def call_api_and_save_json(address, file_name):
    data = requests.get(address)
    with open(file_name, 'w', encoding='utf_8') as file:
        json.dump(data.json(), file, ensure_ascii=False, indent=4)


# 读取json
def read_file(file_name):
    with open(file_name) as file:
        # 转换为字典输出
        return json.load(file)


# 把汇总信息存入excel
def write_info_to_excel(input_data, file_name):
    # 初始化excel文件
    excel_file = Workbook()
    # 指定要存取的sheet
    sheet = excel_file.active
    # 写入数据
    sheet.cell(1, 1, input_data['provinceName'])
    sheet.cell(1, 2, input_data['confirmedCount'])
    sheet.cell(1, 3, input_data['curedCount'])
    sheet.cell(1, 4, input_data['deadCount'])
    # 保存excel
    excel_file.save(file_name)


# 把地区详细信息也存入excel
# 加上汇总信息,加上一个平均值
def write_details_to_excel(input_data, file_name):
    # 初始化EXCEL文件
    excel = Workbook()
    # 创建一个新的details sheet
    sheet = excel.create_sheet('details')
    # 初始化我们的汇总值
    sum_info = 0
    # 取出每一条数据存入
    for i, info in enumerate(input_data):
        sheet.cell(i + 1, 1, info['cityName'])
        sheet.cell(i + 1, 2, info['confirmedCount'])
        sum_info += info['confirmedCount']
        sheet.cell(i + 1, 3, info['curedCount'])
        sheet.cell(i + 1, 4, info['deadCount'])
    # 保存汇总信息
    sheet.cell(len(input_data) + 1, 1, "汇总")
    sheet.cell(len(input_data) + 1, 2, sum_info)
    sheet.cell(len(input_data) + 2, 1, "平均值")
    sheet.cell(len(input_data) + 2, 2, sum_info / len(input_data))
    # 保存excel
    excel.save(file_name)


def write_details_to_excel_with_pandas(input_data, file_name):
    # 利用pandas转换原有的字典数据变成 DataFrame
    df = pd.DataFrame(input_data, columns=['cityName', 'currentConfirmedCount', 'confirmedCount',
                                           'curedCount', 'deadCount', 'locationId', 'cityEnglishName'])
    # 求和
    sum_info = df.sum()
    # 重命名
    sum_info = sum_info.rename('汇总')
    # 将城市名称和位置ID的汇总值变成空
    sum_info['cityName'] = sum_info['locationId'] = ""
    # 取平均值
    avg_info = df.mean()
    # 重命名
    avg_info = avg_info.rename('平均值')
    # 将位置ID的平均值设置为空
    avg_info['locationId'] = ""

    # 中位数
    median_info = df.median()
    median_info = median_info.rename('中位数')
    median_info['locationId'] = ""

    # 添加汇总和平均值信息到DF
    df = df.append(sum_info)
    df = df.append(avg_info)
    df = df.append(median_info)

    # 保存 excel文件
    df.to_excel(file_name)


# call_api_and_save_json('https://lab.isaaclin.cn/nCoV/api/area?province=上海市', '2019_conv.json')
data = read_file('2019_conv.json')
infos = data['results'][0]
# # write_info_to_excel(infos, 'conv.xlsx')
details_info = infos['cities']
# # write_details_to_excel(details_info, 'conv_details.xlsx')
write_details_to_excel_with_pandas(details_info, 'conv_details_pandas.xlsx')
