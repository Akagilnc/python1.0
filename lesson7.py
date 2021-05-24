from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# 定义全局的格式
font_body = Font(name='黑体', size=12, bold=False, color='404040')
alignment_left = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
bd1 = Side(border_style='mediumDashDot', color='000000')
bd2 = Side(border_style='thin', color='000000')
border = Border(left=bd2, right=bd2, top=bd1, bottom=bd1)

# 我要定義title的相關樣式
font_title = Font(name='黑体', bold=True, size=18, color='FFFFFF')
fill_title = PatternFill(fill_type='solid', fgColor='0F4C81')
alignment_title = Alignment(horizontal='center', vertical='center', wrap_text=False)

# 定义重要内容的样式
font_imp = Font('fira', size=16, bold=True, color='000000')
font_imp_red = Font('fira', size=16, bold=True, color='ff0000')
fill_imp = PatternFill(fill_type='solid', fgColor='F3D5AD')

wb = load_workbook('./excel_files/report_fixed.xlsx')

for sheet in wb.worksheets:
    sheet.sheet_view.zoomScale = 300
    # 设置全局的样式
    for row in sheet.rows:
        for cell in row:
            cell.font = font_body
            cell.alignment = alignment_left
            cell.border = border

    # 第一列的样式
    for cell in sheet[1]:
        cell.font = font_title
        cell.fill = fill_title
        cell.alignment = alignment_title

    # 设置列宽以及title的行高
    sheet.column_dimensions['E'].width = 25
    sheet.row_dimensions[1].height = 30
    for col_name in 'ABCDFG':
        sheet.column_dimensions[col_name].width = 15

    # 设置累计确诊人数的样式
    for row_num in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_num].height = 20

        cell_confirmed = sheet['E{}'.format(row_num)]
        cell_confirmed.font = font_imp
        cell_confirmed.fill = fill_imp
        cell_confirmed.number_format = '#,##0'

        if sheet.cell(column=5, row=row_num).value > 10000:
            sheet['C{}'.format(row_num)].font = font_imp_red


wb.save('./excel_files/report_formatted.xlsx')
